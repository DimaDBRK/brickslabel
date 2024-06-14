import openpyxl

def get_info_by_id(file_path, search_id, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if str(row[0]).strip() == str(search_id).strip():
            return row

    return None

def get_all_ids(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    ids = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        ids.append(row[0])
    return ids
